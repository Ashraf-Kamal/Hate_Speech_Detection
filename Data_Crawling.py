# -*- coding: utf-8 -*-
# encoding=utf8
import importlib
import sys
importlib.reload(sys)
#sys.setdefaultencoding('utf8')
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
#from openpyxl.compat import range
import pandas as pd
import os
import Data_Preprocessing as DPF

count=0 # To count no. of text

wb = load_workbook(r"Raw_Datasets\Gao_17\Gao_Final_Raw_Dataset.xlsx")
sheet = wb["Sheet1"]
rows = sheet.max_row
column = sheet.max_column
rows=rows+1
texts=[] 
texts_label=[]   
    # Append texts fetch after pre-processing.
for i in range(2, rows):
  #Text_Id=(sheet.cell(row=i, column=1).value)
  text=(sheet.cell(row=i, column=1).value)      
  label=(sheet.cell(row=i, column=2).value)
  #print(text) 
  if text!=None and len(text.split())>3:
    text_processed=DPF.Preprocessing_Start(text)
    if text_processed!=None and len(text_processed.split())>3:
      texts.append(text_processed)
      texts_label.append(label)
  count +=1
  print (count)
  rows=rows+1  

print (count)
df = {'Text': texts, 'Label': texts_label}
tweets=pd.DataFrame(df)

tweets.to_csv(r"Pre_Processed_Data_Sets\Preprocess_CSV\Gao_17\Gao_Final_Raw_Dataset.csv", sep=',', index=False, encoding='utf-8')

# #Below code is to convert CSV to Excel.
df_new = pd.read_csv(r"Pre_Processed_Data_Sets\Preprocess_CSV\Gao_17\Gao_Final_Raw_Dataset.csv")
writer = pd.ExcelWriter(r"Pre_Processed_Data_Sets\Preprocess_Excel\Gao_17\Gao_Final_Raw_Dataset.xlsx")
df_new.to_excel(writer, index = False)
writer.save()
 
print ('Done!')

#Below code converts excel to csv using dataframe

##df = pd.read_excel("Raw_Datasets\Davison_17\Davidson_Final_Raw_Dataset.xlsx")
##df.to_csv ("Raw_Datasets\Davison_17\Davidson_Final_Raw_Dataset.csv", index = None,  header=True)

#Below code converts csv to excel using dataframe
##read_file = pd.read_csv (r'Raw_Datasets\Davison_17\Davidson_Final_Raw_Dataset.csv')
##read_file.to_excel (r'Raw_Datasets\Davison_17\Davidson_Final_Raw_Dataset.xlsx', index = None, header=True)
